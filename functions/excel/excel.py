"""
для работы с эксель
"""

from datetime import datetime

import openpyxl

from data_class.data import Data
# TODO use Data
def load_excel_book(path: str) -> openpyxl.workbook.workbook.Workbook:
    """
    Загружает книгу Excel из указанного пути.
    
    Args:
    path: строка - путь к файлу Excel.
    
    Returns:
    openpyxl.workbook.workbook.Workbook: загруженная книга Excel.
    
    Raises:
    FileNotFoundError: если файл не найден.
    InvalidFileException: если файл не является допустимым файлом Excel.
    """
    return openpyxl.load_workbook(path)

def get_excel_sheet(
        book: openpyxl.workbook.workbook.Workbook,
        worksheets:int = 0
    ) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Заглушка
    """
    return book.worksheets[worksheets]

def extract_data_from_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[tuple[int, str, bool], tuple[float]]:
    """
    Заглушка
    """
    data = {}
    index_column = 1
    for column in sheet.iter_cols(min_row=1, max_col=sheet.max_column, values_only=True):
        index_column += 1
        _column_values = list(column)
        _list = list((filter(lambda x: x is not None, _column_values)))
        if _list:
            if "Результат" in  _list:
                result =  _list[_list.index("Результат")+1:]
                datalist =  _list[1: _list.index("Результат")]
                _datalist = to_float_list(datalist)
                data[index_column,  _list[0]] = [_datalist, result, False]
            else:
                datalist =  _list[1:]
                _datalist = to_float_list(datalist)
                data[index_column,  _list[0]] = [_datalist, [], False]
    return data

def to_float_list(_list: list) -> list[float]:
    """
    Заглушка
    """
    new_list = []
    for i, k in enumerate(_list):
        try:
            new_list.append((i, float(k)))
        except ValueError as err:
            #допилить что делать если в данных есть ошибка
            print(err)

    return new_list

def different_dict(dict1, dict2):
    """
    Заглушка
    """
    differences = {}

    # Перебираем ключи первого словаря
    for key in dict1.keys():
        if key not in dict2 or dict1[key] != dict2[key]:
            differences[key] = dict1[key]

    # Перебираем ключи второго словаря
    for key in dict2.keys():
        if key not in dict1:
            differences[key] = dict2[key]

    return differences

def read_file_excel(path: str) -> dict[tuple[int, str, bool], tuple[float]]:
    """
    Заглушка
    """
    book = load_excel_book(path)
    sheet = get_excel_sheet(book)
    data = extract_data_from_sheet(sheet)
    if data:
        print(data)
        return Data(data)
    return None


def write_to_file(diff_dict: dict, new_data: dict, sheet: openpyxl.worksheet.worksheet.Worksheet):
    """
    Заглушка
    """
    for key in diff_dict:
        row = 1

        try:
            sheet.cell(row=row, column=key[0]).value = key[1]
            for i, value in enumerate(new_data[key][0]):
                sheet.cell(row=row+i+1, column=key[0]).value = value[1]
        except KeyError:
            sheet.cell(row=row, column=key[0]).value = None
            for i, value in enumerate(diff_dict[key][0]):
                sheet.cell(row=row+i+1, column=key[0]).value = None

        # row += len(new_data[key]) + 1
        # if key[-2]:
        #     sheet.cell(row=row, column=key[0]).value = "Измерения"

        #     for i, val in enumerate(key[-1]):
        #         sheet.cell(row=row+i+1, column=key[0]).value = val

def create_new_sheet(book: openpyxl.workbook.workbook.Workbook, path: str):
    """
    заглушка
    """
    if book.sheetnames[-1] == "Информация":
        ws = get_excel_sheet(book, -1)
        ws.cell(1, 1).value = 'Последнее изменение свершено:'
        ws.cell(1, 2).value = str(datetime.now())
    else:
        book.create_sheet("Информация")
        ws = get_excel_sheet(book, -1)
        ws.cell(1, 1).value = 'Последнее изменение свершено:'
        ws.cell(1, 2).value = str(datetime.now())
    book.save(path)

def save_result_calc_excel(path: str, new_data: dict[tuple[int, str, bool, tuple[float,...]], tuple[float]]) -> dict[tuple[int, str, bool], tuple[float]]:
    """
    Доработать функционал с множеством строк мб много листов

    """
    book = load_excel_book(path)
    sheet = get_excel_sheet(book)
    data = read_file_excel(path)
    diff_dict = different_dict(new_data, data)
    write_to_file(diff_dict, new_data, sheet)
    create_new_sheet(book, path)
    return read_file_excel(path)

def create_new_file(path: str, new_dict: dict[tuple[int, str, bool, tuple[float,...]], tuple[float]]) -> None:
    """
    Заглушка
    """
    book = openpyxl.Workbook()
    sheet = get_excel_sheet(book)
    write_to_file(new_dict, new_dict, sheet)
    create_new_sheet(book, path)

def get_name_column(
        number_column: int,
        data: dict[[tuple[int,str]]: list[int|float]]
    )-> tuple:
    """
    Заглушка
    """
    return tuple(data.keys())[number_column - 1]
